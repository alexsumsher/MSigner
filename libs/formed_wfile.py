#!/usr/bin/env python
# -*- coding: utf8


from openpyxl import Workbook, styles
from wfiles import wfile


class fm_wfile(wfile):
    # make some quick&usefull styles built-in

    def innerforms(self, action, optstr, xltype='xlsx'):
        # add form infomation to list[self.forms]
        # built_in some styles
        # merges: string input with(starts with 1): 'cell_lt,cell_rb;cell_lt,cell_rb...'
        # widths: list of columns(starts with 1): A:100,C:100,D-F:120,J:100
        # heights: list of rows(starts with 1): 1:30,3:30,4-8:35,10:40
        # fixhead[head row frozen]: True
        # frames[thin,left-up-right-bottom]: cell_lt,cell_rb;cell_lt,cell_rb... => extend:
        # cell_lt,cell_rb,rgba
        merges_spliter = lambda s: [[x.strip() for x in cs.split(',')] for cs in s.split(';')]
        w_h_spliter = lambda s: [cs.split(':') for cs in s.split(',')]
        #frames_spliter = lambda s: []
        if action == 'clear':
            self.forms.clear()
            return
        if action == 'merges':
            if 'merges' not in self.forms:
                self.forms['merges'] = []
            self.forms['merges'].extend(merges_spliter(optstr))
        elif action == 'fixhead':
            self.forms['fixhead'] = True
        elif action == 'widths':
            if 'widths' not in self.forms:
                self.forms['widths'] = []
            self.forms['widths'].extend(w_h_spliter(optstr))
        elif action == 'heights':
            if 'heights' not in self.forms:
                self.forms['heights'] = []
            self.forms['heights'].extend(w_h_spliter(optstr))
        elif action == 'frames':
            if 'frames' not in self.forms:
                self.forms['frames'] = []
            self.forms['frames'].extend(merges_spliter(optstr))

    def _innerforms(self, sheet):
        # eval the innerforms to sheet
        # steps: merge-width-height-frame-fixhead
        merges = self.forms.get('merges')
        widths = self.forms.get('widths')
        heights = self.forms.get('heights')
        frames = self.forms.get('frames')
        def _keyuper(fromc, toc):
            # 'A' -> 'ZZ' etc.
            # fool way
            fromc = fromc.upper()
            toc = toc.upper()
            line = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            emptybit = ' '
            maxi = len(line) - 1
            bits = [x for x in fromc]
            bits.reverse()
            if len(bits) < len(toc):
                for _ in xrange(len(toc)-len(bits)):
                    bits.append(emptybit)
            bit_inc = True
            out = ""
            while 1:
                out = ''.join(bits)[::-1].strip()
                yield out
                if out == toc:
                    break
                bit_inc = True
                cbit = 0
                for b in bits:
                    if bit_inc is False:
                        break
                    i = line.find(b)
                    if i == maxi:
                        bits[cbit] = line[0]
                        bit_inc = True
                    elif i == -1 and bits[cbit] == emptybit:
                        # empty bit as space
                        bits[cbit] = line[0]
                        bit_inc = False
                    else:
                        bits[cbit] = line[i+1]
                        bit_inc = False
                    cbit += 1
        def _csplit(s):
            for _ in xrange(len(s)):
                if s[_].isdigit():
                    return s[:_],int(s[_:])
        def _frame2cells(lt, rb):
            # lt="A1", rt="B3" etc.
            lt_n,lt_v = _csplit(lt)
            rb_n,rb_v = _csplit(rb)
            for c in _keyuper(lt_n, rb_n):
                for v in xrange(lt_v, rb_v+1):
                    yield c + str(v)
        if widths:
            for c,w in widths:
                w = int(w.strip())
                c = c.strip().upper()
                if '-' in c:
                    f,t = c.split('-')
                    for _ in _keyuper(f, t):
                        sheet.column_dimensions[_].width = w
                else:
                    sheet.column_dimensions[c].width = w
        if heights:
            for r,h in heights:
                h = int(h.strip())
                if '-' in r:
                    s,e = r.split('-')
                    for _ in xrange(int(s.strip()), int(e.strip())+1):
                        sheet.row_dimensions[_].height = h
                else:
                    sheet.row_dimensions[int(r)].height = h
        if merges:
            for corner in merges:
                #for lt,rb in merges:
                #   sheet.merge_cells('%s:%s' % (lt, rb))
                sheet.merge_cells(':'.join(corner))
        if frames:
            thin = styles.Side(border_style="thin", color="000000")
            for frame in frames:
                if len(frame) == 3:
                    rgba = frame.pop()
                else:
                    rgba = None
                for c in _frame2cells(*frame):
                    sheet[c].border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    if rgba:
                        sheet[c].fill = styles.PatternFill('solid', fgColor=rgba)
        if self.forms.get('fixhead'):
            sheet.freeze_panes = 'A2'
        return sheet

    def _infm_check(self, action, fullcheck=False):
        # check if the optdata in forms are correct
        return True

    def do_export(self, filename='', startrow=0, startcol=0):
        xltype = 'xlsx'
        filename = (filename or str(uuid.uuid1())) + '.' + xltype
        self.export_path = self._checkfile(filename, xltype=xltype)
        return self._exp_xls(xltype=xltype, startrow=startrow, startcol=startcol)


if __name__ == '__main__':
    dict_input = [{'name': '张三', 'age': 13, 'gender': 'boy', 'date': '0619'}, {'name': '李四', 'age': 14, 'gender': 'girl', 'date': '0829'}, {'name': '王五', 'age': 23, 'gender': 'women', 'date': '1011'}, {'name': '赵六', 'age': 22, 'gender': 'man', 'date': '0609'}]
    getter = ['age', 'name']
    head = ['NAME', 'AGE', 'GENDER']
    w = fm_wfile(dict_input, datagetter=getter, headrow=head, folder='d:\\tmp2')
    w.innerforms('merges', 'A10,F11')
    w.innerforms('widths','A-F:35')
    w.innerforms('heights', '1:30,2-10:20')
    w.innerforms('frames', 'A1,C1,FCD5B400')
    #print w.forms
    print(w.forms)
    w.do_export('tt1', startrow=2)
