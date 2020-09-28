#!/usr/bin/env python
# -*- coding: utf8
# 
# 基于openpyxl的excel文件读取、输出模块
# wfile：常用，导出xls,xlsx,csv，带自定义部分内容，用于导出单sheet文件
# wfile_multi：常用，基于wfile，用于导出多sheet文件
# 另外文件formed_wfile为增加常用的格式效果
# 
import os
import sys
import shutil
import csv
import xlrd
import xlwt
# for template mode, we have to use openpyxl with xlsx
from openpyxl import load_workbook, Workbook
import uuid
import time
import logging

from localdata import binfile

logger = logging.getLogger(__name__) or logging

# MODULE
# use xlrd/xlwt and openpyxl to handle xls/xlsx files;;

class xltpl(object):
    # a template object for excel
    # a template def by 2 file: exel file + defination file(.def)
    workfd = ''
    EXT = 'def'

    @classmethod
    def gen_by_xlsx(cls, xfilename, withdesc=False):
        # generate a def from xlsxfile, the cells with value like: "$$keyname$description"
        if not os.path.exists(xfilename):
            xfilepath = os.path.join(cls.workfd, xfilename)
            if not os.path.exists(xfilepath):
                raise ValueError("NOT found xlsx file: %s" % xfilepath)
        else:
            xfilepath = xfilename
            xfilename = os.path.split(xfilepath)[1]
        xfile = load_workbook(xfilepath)
        # for cells merging we have to formatting_info=True
        tbl = xfile.worksheets[0]
        export = []
        for r in tbl.iter_rows():
            for c in r:
                v = c.value
                if v and v.startswith('$$'):
                    splits = v.split('$')[2:]
                    slen = len(splits)
                    # '$$key$desc' => ['', '', key, desc]
                    # '$$key' => ['', '', key]
                    # export: [key, desc, coord]
                    if slen == 1:
                        splits.append(u'')
                        splits.append(c.coordinate)
                    elif slen == 2:
                        splits.append(c.coordinate)
                    else:
                        logger.warning("NOT Correct IMPORT: %s" % v)
                    export.append(splits)
        return cls.add(xfilename, export)

    @classmethod
    def add(cls, xlsxfilename, defdata):
        # defdata a data object
        if not os.path.exists(os.path.join(cls.workfd, xlsxfilename)):
            raise ValueError("XLS FILE: %s NOT FOUND IN %s!" % (xlsxfilename, cls.workfd))
        if not isinstance(defdata, list):
            raise ValueError("not list")
        deffile = os.path.splitext(xlsxfilename)[0] + '.' + cls.EXT
        ofile = binfile(cls.workfd)
        return ofile.save_data(defdata, deffile)

    @classmethod
    def collect(cls, folder=None):
        folder = folder or cls.workfd
        singles = set()
        tpls = set()
        for f in os.listdir(folder):
            f = f.os.path.splitext(f)[0]
            if f in singles:
                singles.pop(f)
                tpls.add(f)
        return list(tpls)
    
    def __init__(self, name, folder=None):
        folder = folder or self.__class__.workfd
        xlsxpath = os.path.join(folder, name + '.xlsx')
        defpath = os.path.join(folder, name + '.' + self.__class__.EXT)
        if not os.path.exists(xlsxpath):
            raise ValueError("NOT FOUND TEMPLATE!")
        if not os.path.exists(defpath):
            self.__class__.gen_by_xlsx(xlsxpath)
        if os.path.exists(defpath):
            self.name = name
            self.folder = folder
            self.__load_bindef()
        else:
            raise ValueError("no defdata & not able to generate defdata!")

    @property
    def tplfname(self):
        return self.name + '.xlsx'

    @property
    def deffname(self):
        return self.name + '.' + self.__class__.EXT

    def __load_bindef(self):
        self.defdata = binfile(self.folder).load_data(self.deffname)

    def update(self, defdata):
        if isinstance(defdata, list):
            self.defdata = defdata
            deffile = self.deffname
            return binfile(self.folder).save_data(self.defdata, deffile)

    def export(self, emode=0):
        self.__load_bindef()
        # [key, desc, coord]
        # emode: 0 => all; 1=> key only; 2=> key & desc
        if emode == 1:
            return [_[0] for _ in self.defdata]
        elif emode == 2:
            return [_[:-1] for _ in self.defdata]
        else:
            return self.defdata

    # Update: 2019-11-10
    # part of export xlsx from tmplate+def
    def _key_datas(self, xobj, datas, defs=None):
        defs = defs or self.defdata
        if isinstance(defs, dict):
            defs = [[_['keyname'], _.get('description'), _['coordinate']] for _ in defs]
        for r in defs:
            xobj[r[2]] = datas.get(r[0], "")
        return True

    def _row_datas(self, xobj, datas):
        # {row_start: (startsfrom1), col_start(starsfrom1), rows: [[row1], [row2], ...]}
        size = len(datas['rows'])
        #xobj.insert_rows(datas['row_start'], amount=size)
        r = datas['row_start']
        for row in datas['rows']:
            c = datas.get('col_start', 1)
            for _ in range(len(row)):
                xobj.cell(r, c+_, row[_])
            r += 1
        return True

    def make_file(self, fname, keydatas, rowdatas=None, expdir=None, overwrite=True):
        if not self.defdata or len(self.defdata) == 0:
            logger.warning("No defdata!")
            return
        expdir = expdir or self.workfd
        if not fname.endswith('xlsx'):
            fname = fname + '.xlsx'
        fname = os.path.join(expdir, fname)
        if os.path.exists(fname) and not overwrite:
            raise ValueError("export target(%s) exists!" % fname)
        logger.info('xltpl->%s' % fname)
        xfile = load_workbook(os.path.join(self.folder, self.tplfname))
        ws = xfile.active
        self._key_datas(ws, keydatas)
        if rowdatas:
            self._row_datas(ws, rowdatas)
        """
        def data:
        [['title', '南油小学20   ——20   学年度  学期期末考试试卷分析', 'A1'], 
        ['subject', '科目', 'B2'], 
        ['teacher', '教师', 'D2'], 
        ['classsic', '典型题', 'B12'], 
        ['totally', '总体概述', 'B13'], 
        ['problems', '存在问题', 'B14'], 
        ['improve', '改进 措施', 'B15']]
        keydatas:
        {'keyname': value, ..., }
        rowdatas:
        {row_start: (startsfrom1), col_start(starsfrom1), rows: [[row1], [row2], ...]}
        """
        xfile.save(filename=fname)
        return fname



class wfile(object):
    """
    生成excel数据文件，基于单sheet
    cls.set_folder：设定默认输出文件夹；
    接受三种数据格式：\t,\n分割文本；list-list；list-dict
    wfile(...): 
    headrow: 标题行，list，一行内容；headfirst(boolean)是否数据第一行为标题行；
    folder：输出文件夹；
    calculater：计算器，函数；对每行数据进行预处理，直接修改其内容，不返回；
    datagetter：数据提取，对每行数据（list of value 或者 list of dict）进行提取，对于输入数据为list of dict，datagetter应当是list of keys；对于list of list，则是list of int（以0开始的数组标号）

    free_cells:设定补充数据，对于非标准数据的补充，可以是以1开始的一行，或者一列，或者row,col；可以通过该函数设置多个补充数据，注意有前后顺序; 如果该函数第一个参数（模式）为-1则清空freecellls
    """
    workfd = ''
    allow_types = ['csv', 'xls', 'xlsx']

    @classmethod
    def set_folder(cls, fd):
        assert os.path.isfolder(fd) and os.path.exists(fd)
        self.workfd = fd

    @classmethod
    def from_tpl(cls, template, values, filename=None, exportdir=None):
        # template: template name:
        # [[key, desc, coord]] => key => datastream: {key:value, key:value...}
        tpler = xltpl(template, cls.workfd)
        tplfile = os.path.join(cls.workfd, template + '.xlsx')
        outfile = os.path.join(exportdir or cls.workfd, (filename or str(uuid.uuid1())) + '.xlsx')
        w = load_workbook(tplfile)
        s = w.worksheets[0]
        for key,desc,coord in tpler.export():
            v = values.get(key)
            if v:
                s[coord].value = v
        w.save(filename=outfile)

    @classmethod
    def complex_tpl(cls, tplfile, export_name, form_datas, exportdir=None):
        # get tpl and export with formdata like:
        # []
        pass

    def __init__(self, datalist, datagetter=None, headrow=None, headfirst=False, folder="", calculater=None, replace=None):
        if isinstance(datalist, (str, unicode)):
            self.datalist = self._serial(datalist)
        else:
            self.datalist = datalist
        if isinstance(self.datalist, (list,tuple)) and isinstance(self.datalist[0], dict):
            if not datagetter:
                raise ValueError("we got a dictionary of row-data, a datagetter is needed.")
            self.datagetter = self._dict2list(datagetter, replace)
        elif isinstance(datagetter, list):
            # callable or None
            self.datagetter = self._list2list(datagetter, replace)
        else:
            self.datagetter = None
        # check for datamode: strign,dict,list/tuple[single verion not surpport iterator]
        if hasattr(self.datalist, 'next'):
            headfirst = False
        if headfirst is True:
            self.headrow = self.datalist.pop(0)
        elif headrow:
            self.headrow = headrow
        else:
            self.headrow = None
        if callable(calculater):
            self.calculater = calculater
        else:
            self.calculater = None
        self.fname = None
        self.export_path = None
        self.forms = {} # not using, keep for formedwfile
        self.freecells = []
        self.workfd = folder or self.__class__.workfd or os.getcwd()

    def _dict2list(self, datagetter, replace=None):
        #print('replace',replace)
        def convert(dictin):
            outlist = []
            for _ in datagetter:
                if replace:
                    v = dictin.get(_)
                    v = None if v == replace else v
                    outlist.append(v)
                else:
                    outlist.append(dictin.get(_))
            return outlist
        return convert

    def _list2list(self, datagetter, replace=None):
        def convert(listin):
            outlist = []
            for _ in datagetter:
                v = listin[_]
                if replace:
                    v = None if v == replace else v
                outlist.append(v)
            return outlist
        return convert

    def _serial(self, datastream, cell_split='\t', row_split='\n'):
        # datalist is string => list of list
        cell_split = cell_split or '\t'
        row_split = row_split or '\n'
        rows = datastream.split(row_split)
        datalist = []
        for r in rows:
            datalist.append(r.split(cell_split))
        return datalist

    def _checkfile(self, checkfile, xltype='xls'):
        # checkfile(could be fullpath or just filename) whether exists in workfolder
        folder,filename = os.path.split(checkfile)
        if folder and not os.path.exists(checkfile):
            return checkfile
        else:
            folder = folder or self.workfd
        p = filename.rfind('.')
        if p > 0:
            fnbody = filename[:p]
            fnext = filename[p:]
        else:
            fnbody = filename
            fnext = '.' + xltype
        fpth = fnbody + fnext
        while True:
            epth = os.path.join(folder, fpth)
            if os.path.exists(epth):
                fpth = fnbody + '_' + str(time.time())[:-3] + fnext
            else:
                break
        self.fname = fpth
        return epth

    def _xls_copy(self, fromfile, newfile='', xltype='xls'):
        # copy template xls file to a new one
        xltype = xltype.lower()
        if not fromfile.endswith(xltype):
            fromfile = fromfile + '.' + xltype
        if not os.path.exists(fromfile):
            fromfile2 = os.path.join(self.workfd, fromfile)
            if not os.path.exists(fromfile2):
                raise ValueError("%s not found!" % fromfile)
        newfile = newfile or os.path.split(fromfile)[1]
        if not newfile.endswith(xltype):
            newfile = newfile + '.' + xltype
        newfile = self._checkfile(newfile, xltype=xltype)
        shutil.copyfile(fromfile, newfile)
        return newfile

    def innerforms(self, action, *args):
        # not implemented in base object
        raise NotImplementedError()

    def _innerforms(self, sheet):
        # not implemented in base object
        raise NotImplementedError()

    def free_cells(self, mode, *args, **kwargs):
        # freely add cell value
        # work finally
        # self.freecells: [{mode,data[row,col,val] or data[vals]}];
        # mode0: cell(0),row(1),col(2)
        # mode1: cell, data[rownum,colnum,val]
        # mode2: row(1),col(2): number(rownum or colnum), data:[vals list]
        if mode == -1:
            self.freecells = []
            return
        if mode == 0:
            self.freecells.append(dict(mode=0, data=args if args else [kwargs['row'], kwargs['col'], kwargs['val']]))
            #self.freecells.append(dict(mode=0, data=[kwargs['row'], kwargs['col'], kwargs['val']]))
        elif mode == 1:
            self.freecells.append(dict(mode=mode, number=kwargs['row'], start=kwargs['start'] or 1 ,data=kwargs['val']))
        elif mode == 2:
            self.freecells.append(dict(mode=mode, number=kwargs['col'], start=kwargs['start'] or 1, data=kwargs['val']))

    def freecell_solve(self, sheet, xltype='xlsx'):
        # addtion
        # when row and number==-1, append to final
        if xltype == 'xlsx':
            writer = lambda r,c,v:sheet.cell(row=r,column=c,value=v)
            isxlx = True
        else:
            writer = sheet.write
            isxls = False
        for line in self.freecells:
            data = line['data']
            mode = line['mode']
            if mode == 0:
                writer(data[0], data[1], data[2])
            elif mode == 1:
                start = line['start'] if isxlx else line['start'] - 1
                row = line['number']
                if row == -1:
                    row = sheet.max_row + 1
                for _ in xrange(len(data)):
                    writer(row, start+_, data[_])
            elif mode == 2:
                start = line['start'] if isxlx else line['start'] - 1
                col = line['number']
                if col == -1:
                    col = sheet.max_column + 1
                for _ in xrange(len(data)):
                    writer(start+_, col, data[_])

    def template_export(self, tpfilepath, newfilename, startrow=0, startcol=0):
        # works with a template, to which fill with datas
        # in all probability the headrow is exists, just put data row by row...
        wb = load_workbook(tpfilepath)
        self.export_path = self._checkfile(newfilename, xltype='xlsx')
        ws = wb.active
        rowid = 1
        if self.forms:
            ws = self._innerforms(ws)
        if self.headrow:
            # if headrow, skip startrow
            for c in xrange(1, len(self.headrow) + 1):
                ws.cell(row=1, column=c, value=self.headrow[c-1])
                #worksheet.write(0, c, self.headrow[c])
            rowid = 2
        if startrow > 1:
            rowid = startrow
        for row in self.datalist:
            if self.calculater:
                self.calculater(rowdata)
            rowdata = self.datagetter(row) if callable(self.datagetter) else row
            cellid = startcol if startcol > 1 else 1
            for cell in rowdata:
                if cell is None:
                    continue
                ws.cell(row=rowid, column=cellid, value=cell)
                #worksheet.write(rowid, cellid, cell)
                cellid += 1
            rowid += 1
        if self.freecells:
            self.freecell_solve(ws)
        wb.save(filename=self.export_path)

    def do_export(self, filename='', xltype='xlsx', startrow=0, startcol=0, overwrite=False):
        xltype = xltype.lower()
        if xltype not in self.__class__.allow_types:
            raise ValueError("%s: type not supported.")
        filename = (filename or str(uuid.uuid1())) + '.' + xltype
        if overwrite:
            self.export_path = os.path.join(self.workfd, filename)
        else:
            self.export_path = self._checkfile(filename, xltype=xltype)
        if xltype == 'csv':
            return self._exp_csv(startrow=startrow, startcol=startcol)
        else:
            return self._exp_xls(xltype=xltype, startrow=startrow, startcol=startcol)

    def _exp_xls(self, xltype='xls', startrow=0, startcol=0):
        # startrow/col starts with 1
        if xltype == 'xls':
            wbook = xlwt.Workbook(encoding='utf-8')
            worksheet = wbook.add_sheet('sheet1')
            writer = worksheet.write
            startn = 0
            startrow = startrow - 1
        elif xltype == 'xlsx':
            wbook = Workbook()
            worksheet = wbook.active
            writer = lambda r,c,v:worksheet.cell(row=r,column=c,value=v)
            startn = 1
        else:
            raise RuntimeError("%s: unknown file type!" % xltype)
        rowid = startn
        # inner form
        if xltype == 'xlsx' and self.forms:
            worksheet = self._innerforms(worksheet)
        if self.headrow:
            for c in xrange(startn, len(self.headrow) + startn):
                writer(startn, c, self.headrow[c-startn])
                #worksheet.write(0, c, self.headrow[c])
            rowid = startn + 1
        if startrow > 0:
            rowid = startrow
        for row in self.datalist:
            if self.calculater:
                self.calculater(row)
            rowdata = self.datagetter(row) if self.datagetter else list(row)
            cellid = startn + startcol
            for cell in rowdata:
                if cell is None:
                    cellid += 1
                    continue
                writer(rowid, cellid, cell)
                #worksheet.write(rowid, cellid, cell)
                cellid += 1
            rowid += 1
        if self.freecells:
            self.freecell_solve(worksheet, xltype=xltype)
        wbook.save(self.export_path) if xltype == 'xls' else wbook.save(filename=self.export_path)
        return rowid

    def _exp_csv(self, startrow=0, startcol=0):
        rows = 0
        with open(self.export_path, 'wb') as csvfile:
            writer = csv.writer(csvfile)
            if self.headrow:
                writer.writerow([s.encode('utf-8') for s in self.headrow])
                rows = 1
            for row in datalist:
                if self.calculater:
                    self.calculater(row)
                rowdata = self.datagetter(row) if self.datagetter else row
                writer.writerow([s.encode('utf-8') for s in rowdata])
                rows += 1
        return rows


class wfile_multi(wfile):
    # support multi sheet export
    
    def __init__(self, datalist, datagetter=None, headrow=None, folder="", calculater=None, replace=None):
        assert hasattr(datalist, 'gi_frame') or isinstance(datalist[0], (list, tuple, dict)),"iterdata is needed."
        if datagetter and isinstance(datagetter[0], int):
            self.datagetter = self._list2list(datagetter, replace)
        elif datagetter:
            self.datagetter = self._dict2list(datagetter, replace)
        else:
            self.datagetter = None
        if callable(calculater):
            self.calculater = calculater
        else:
            self.calculater = None
        self.index = 0
        self.datalist = datalist
        # check for datamode: strign,dict,list/tuple[single verion not surpport iterator]
        self.forms = None
        self.freecells = None
        self.headrow = headrow
        self.export_path = None
        self.workfd = folder or self.__class__.workfd or os.getcwd()

    def free_cells(self, mode, **kwargs):
        # free cells setting on multi
        if mode == -1:
            self.freecells = []
            return
        if self.freecells is None:
            self.freecells = []
        if mode == 0:
            self.freecells.append(dict(mode=0, data=[kwargs['row'], kwargs['col'], kwargs['val']]))
        elif mode == 1:
            self.freecells.append(dict(mode=mode, number=kwargs['row'], start=kwargs['start'] or 1 ,data=kwargs['val']))
        elif mode == 2:
            self.freecells.append(dict(mode=mode, number=kwargs['col'], start=kwargs['start'] or 1, data=kwargs['val']))
        elif isinstance(mode, (str, unicode)):
            # mode other: 'sheetname';
            # set value to special sheet by title
            self.freecells.append(dict(mode=mode, data=[kwargs['row'], kwargs['col'], kwargs['val']]))
        else:
            logger.warning("Unknown free cell mode!")

    def freecell_solve(self, sheet, xltype='xlsx'):
        # addtion make content freely
        # when row and number==-1, append to final
        if xltype == 'xlsx':
            writer = lambda r,c,v:sheet.cell(row=r,column=c,value=v)
            isxlx = True
        else:
            writer = sheet.write
            isxls = False
        for line in self.freecells:
            data = line['data']
            mode = line['mode']
            if mode == 0:
                writer(data[0], data[1], data[2])
            elif mode == 1:
                start = line['start'] if isxlx else line['start'] - 1
                row = line['number']
                if row == -1:
                    row = sheet.max_row + 1
                for _ in xrange(len(data)):
                    writer(row, start+_, data[_])
            elif mode == 2:
                start = line['start'] if isxlx else line['start'] - 1
                col = line['number']
                if col == -1:
                    col = sheet.max_column + 1
                for _ in xrange(len(data)):
                    writer(start+_, col, data[_])
            elif mode == sheet.title:
                writer(data[0], data[1], data[2])

    def multi_export_xls(self, filename='', xltype='xlsx', sheetnames=None, startrow=0, startcol=0, skip_empty=True):
        # startrow/startcol starts with 1
        xltype = xltype.lower()
        if xltype not in self.__class__.allow_types:
            raise ValueError("%s: type not supported.")
        filename = (filename or str(uuid.uuid1())) + '.' + xltype
        #filename = self._checkfile(filename, xltype=xltype)
        # if startrows
        if xltype == 'csv':
            filename = filename.replace(".csv", "[%s].csv")
        elif xltype == 'xls':
            wbook = xlwt.Workbook(encoding='utf-8')
            sheeter = wbook.add_sheet
            saver = wbook.save
            startn = 0
            startrow = startrow - 1
        elif xltype == 'xlsx':
            wbook = Workbook()
            #wbook.remove_sheet(wbook.get_active_sheet())
            wbook.remove(wbook.get_active_sheet())
            sheeter = lambda t:wbook.create_sheet(title=t)
            saver = lambda f:wbook.save(filename=f)
            startn = 1
        export_path = os.path.join(self.workfd, filename)
        self.export_path = export_path
        self.index = 0
        for datalist in self.datalist:
            if len(datalist) == 0 and skip_empty:
                self.index += 1
                continue
            if xltype == 'csv':
                self.export_path = export_path % self.index
                self._exp_csv()
            else:
                self.sheetname = sheetnames[self.index] if isinstance(sheetnames, list) else ('sheet%s' % self.index+1)
                worksheet = sheeter(self.sheetname)
                if self.forms:
                    worksheet = self._innerforms(worksheet)
                writer = worksheet.write if hasattr(worksheet, 'write') else lambda r,c,v:worksheet.cell(row=r,column=c,value=v)
                rowid = startn
                if self.headrow:
                    for c in xrange(startn, len(self.headrow) + startn):
                        writer(startn, c, self.headrow[c - startn])
                    rowid = startn + 1
                if startrow > 0:
                    rowid = startrow
                for row in datalist:
                    if self.calculater:
                        self.calculater(row)
                    rowdata = self.datagetter(row) if self.datagetter else row
                    cellid = startn  + startcol
                    for cell in rowdata:
                        if cell is None:
                            continue
                        writer(rowid, cellid, cell)
                        cellid += 1
                    rowid += 1
                if self.freecells:
                    self.freecell_solve(worksheet, xltype=xltype)
            self.index += 1
        if xltype != 'csv':
            saver(export_path)
        return self.index

    def template_export(self, tpfilepath, newfilename, sheetnames=None, startrow=1, startcol=1, skip_empty=True, encode=False):
        # works with a template, to which fill with datas
        # in all probability the headrow is exists, just put data row by row...
        if not os.path.exists(tpfilepath):
            tpfilepath = os.path.join(self.workfd, tpfilepath)
            if not os.path.exists(tpfilepath):
                raise ValueError("file not found.")
        wb = load_workbook(tpfilepath)
        self.export_path = self._checkfile(newfilename, xltype='xlsx')
        wsa = wb.active
        if self.forms:
            wsa = self._innerforms(wsa)
        self.index = 0
        for datalist in self.datalist:
            if len(datalist) == 0 and skip_empty:
                self.index += 1
                continue
            ws = wb.copy_worksheet(wsa)
            ws.title = sheetnames[self.index] if isinstance(sheetnames, list) else ('sheet%s' % (self.index + 1))
            rowid = 1
            if self.headrow:
                for c in xrange(1, len(self.headrow) + 1):
                    ws.cell(row=1, column=c, value=self.headrow[c-1])
                rowid = 2
            if startrow >= 2:
                rowid = startrow
            if startrow <= 1:
                for row in datalist:
                    if self.calculater:
                        self.calculater(row)
                    ws.append(self.datagetter(row) if self.datagetter else row)
            else:
                for row in datalist:
                    if self.calculater:
                        self.calculater(row)
                    rowdata = self.datagetter(row) if self.datagetter else row
                    cellid = startcol
                    for cell in rowdata:
                        if cell is None:
                            continue
                        #if encode and isinstance(cell, (str, unicode)):
                        #    cell = cell.encode('gbk')
                        ws.cell(row=rowid, column=cellid, value=cell)
                        #worksheet.write(rowid, cellid, cell)
                        cellid += 1
                    rowid += 1
            if self.freecells:
                self.freecell_solve(ws)
            self.index += 1
        #wb.remove_sheet(wsa)
        wb.remove(wsa)
        wb.save(filename=self.export_path)


class xlreader(object):
    # read excelfile(xls, xlsx); csv is not support yet
    # work as a common interface[openxlxs, xlrd]
    # export data
    
    @classmethod
    def scope(cls, filepath):
        pass

    def __init__(self, filepath, with_header=False):
        if os.path.exists(filepath):
            self.filepath = filepath
        else:
            raise ValueError("file: %s not found!" % filepath)
        self._head = None
        self._data = None
        self.with_header = with_header
        self.xfile = None
        self._ftype(filepath)

    def _ftype(self, fpth):
        ext = fpth.split('.')[-1].upper()
        if ext == 'XLS':
            self.type = 0
        elif ext == 'XLSX':
            self.type = 1
        else:
            raise ValueError("illegle file type!")

    def _open(self, sheetn=0):
        if self.type == 0:
            self.xfile = xlrd.open_workbook(self.filepath) if not self.xfile else self.xfile
            return self.xfile.sheets()[sheetn]
        else:
            self.xfile = load_workbook(self.filepath, read_only=True) if not self.xfile else self.xfile
            return self.xfile.worksheets[sheetn]

    def _getrows(self, row_start=1, row_len=0, colist=None):
        # colist: (1,2,3...) starts with 1!
        tbl = self._open()
        if row_start < 1:
            return None
        if self.with_header and row_start == 1:
            row_start = 2
        export = []
        if self.type == 0:
            x = 0
            row_len = tbl.nrows if row_len <= 0 else min(tbl.nrows, row_len)
            row_end = row_start + row_len
            for r in tbl.get_rows():
                x += 1
                if x < row_start:
                    continue
                elif x >= row_end:
                    break
                export.append([c.value for c in r] if not colist else [r[i-1].value for i in colist])
        else:
            if row_len <= 0:
                rows = tbl[row_start:tbl.max_row]
            else:
                row_end = min(tbl.max_row, row_start + row_len)
                rows = tbl[row_start:row_end]
            for r in rows:
                export.append([c.value for c in r] if not colist else [r[i-1].value for i in colist])
        return export

    def iterrows(self, formater=None, colist=None):
        if self.type == 0:
            self.iter = self._open().get_rows()
        else:
            self.iter = self._open().rows
        if self.with_header:
            self.iter.next()
        for r in self.iter:
            if formater:
                yield formater([c.value for c in r] if not colist else [r[i-1].value for i in colist])
            else:
                yield [c.value for c in r] if not colist else [r[i-1].value for i in colist]

    @property
    def data(self):
        if self.data:
            return self.data
        self.data = self._getrows()
        return self.data

    @property
    def head(self):
        if not self.with_header:
            return None
        if self._head:
            return self._head
        tbl = self._open()
        if self.type == 0:
            row = tbl.row(0)
            self._head = [c.value for c in row]
        else:
            row = tbl[1]
            self._head = [c.value for c in row]
        return self._head

    def first_line(self):
        linedata = self._getrows(2, 1) if self.with_header else self._getrows(1, 1)
        return linedata[0] if linedata else None
        
    def get_rows(self, rstart, rlen):
        if rstart == rlen == 0:
            return self.data
        return self._getrows(rstart, rlen)


if __name__ == '__main__':
    #xltpl.workfd = 'D:\\pyworks\\fenlu\\xlsx_tpl'
    #xltpl.gen_by_xlsx('ny_cal.xlsx')
    #print xltpl('ny_cal').export()
    #wfile.workfd = 'D:\\pyworks\\fenlu\\xlsx_tpl'
    #testdata = {'title': 'a test title', 'course': '语文？', 'total': 10000, 'a': 20, 'top': 99}
    #wfile.from_tpl('ny_cal', testdata, u'测试文档', 'D:\\pyworks\\fenlu\\tmp')
    wfile_multi.workfd = 'e:\\temps'
    wfile.workfd = 'e:\\temps'
    tpler = 'e:\\temps\\csbkx.xlsx'
    def gender0(rows):
        for r in xrange(rows):
            yield range(10)
    #w = wfile(gender0(100))
    #dict_input = [{'name': '张三', 'age': 13, 'gender': 'boy', 'date': '0619'}, {'name': '李四', 'age': 14, 'gender': 'girl', 'date': '0829'}, {'name': '王五', 'age': 23, 'gender': 'women', 'date': '1011'}, {'name': '赵六', 'age': 22, 'gender': 'man', 'date': '0609'}]
    #head = [u'年龄', u'姓名', u'日期', u'分数']
    #getter = ['age', 'name', 'date']
    #w = wfile(dict_input, datagetter=getter, headrow=head)
    #w.free_cells(1, row=-1, start=0, val=['a', 'b', 'c', 'd', 'e'])
    #w.do_export(u'语文分数表', xltype='xlsx', startrow=2)
    #pass
    def gender(pages):
        for p in range(pages):
            expt = []
            for r in range(1000):
                expt.append(['%d-%d-%d' % (p, r, _) for _ in range(10)])
            yield expt
    def caler(r):
        r[0] = 'head:%s' % r[0]
        r[3] = 'mid: %s' % r[3]
    w = wfile_multi(gender(3), calculater=caler)
    w.template_export(tpler, 'aaa', startrow=2)
    #dict_input = [[{'a':111, 'b':112, 'c':113, 'd': 114}, {'a':121, 'b':122, 'c':123, 'd': 124}, {'a':131, 'b':132, 'c':133, 'd': 134}], [{'a':211, 'b':212, 'c':213, 'd': 214}, {'a':221, 'b':222, 'c':223, 'd': 224}, {'a':231, 'b':232, 'c':233, 'd': 234}]]
    #dict_list = ['a','c', 'd', 'b']
    #w = wfile_multi(dict_input, datagetter=dict_list)
    #w.multi_export_xls(filename='aaa')
    #list_input = [[[111, 112, 113, 114], [121, 122, 123, 124], [131, 132, 133, 134]], [[211, 221, 223, 224], [221, 222, 223, 224], [231, 232, 233, 234]]]
    #w = wfile_multi(list_input)
    #w.multi_export_xls(filename='aac')